from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.models.gain_record import GainRecord


TWO_PLACES = Decimal("0.01")


class ExcelReporter:
    """
    Generates .xlsx report with two sheets:
    - Tax Summary
    - Audit List

    Financial precision strategy:
    - All calculations remain in Decimal (business layer safety)
    - Values are quantized to 2 decimal places for presentation
    - ROUND_HALF_UP is used for regulatory-safe rounding
    - Excel number_format enforces visual consistency
    - No float conversion is performed (prevents precision loss)
    """

    HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    LOSS_FONT = Font(color="FF0000")
    GAIN_FONT = Font(color="008000")

    MONEY_FORMAT = "#,##0.00"

    def generate(
        self,
        summary: dict,
        records: list[GainRecord],
        output_path: str = "output_report.xlsx",
    ) -> str:

        wb = Workbook()

        self._write_summary_sheet(wb, summary)
        self._write_audit_sheet(wb, records)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        path = Path(output_path)
        wb.save(path)

        return str(path)

    # -------------------------------------------------------
    # SUMMARY SHEET
    # -------------------------------------------------------

    def _write_summary_sheet(self, wb: Workbook, summary: dict) -> None:
        ws = wb.create_sheet("Tax Summary")

        ws.append(["Metric", "Value"])

        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        rows = [
            ("Fiscal Year", summary["fiscal_year"]),
            ("Total Gain (TL)", self._round(summary["total_gain_tl"])),
            ("Total Loss (TL)", self._round(summary["total_loss_tl"])),
            ("Net Gain (TL)", self._round(summary["net_gain_tl"])),
            ("Exemption Applied (TL)", self._round(summary["exemption_applied_tl"])),
            ("Taxable Base (TL)", self._round(summary["taxable_base_tl"])),
            ("Estimated Tax (TL)", self._round(summary["estimated_tax_tl"])),
        ]

        for label, value in rows:
            ws.append([label, value])

            if isinstance(value, Decimal):
                cell = ws.cell(row=ws.max_row, column=2)
                self._format_money_cell(cell)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    # -------------------------------------------------------
    # AUDIT SHEET
    # -------------------------------------------------------

    def _write_audit_sheet(self, wb: Workbook, records: list[GainRecord]) -> None:
        ws = wb.create_sheet("Audit List")

        headers = ["Ticker", "Matched Qty", "Realized Gain (TL)", "Indexed", "Audit Note"]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

        for record in records:
            ws.append([
                record.ticker,
                record.matched_qty,  # quantity is not forced to 2 decimals
                self._round(record.realized_gain_tl),
                "Yes" if record.is_indexed else "No",
                record.audit_note,
            ])

            qty_cell = ws.cell(row=ws.max_row, column=2)
            qty_cell.alignment = Alignment(horizontal="right")

            gain_cell = ws.cell(row=ws.max_row, column=3)
            self._format_money_cell(gain_cell)

            gain_cell.font = (
                self.GAIN_FONT if record.realized_gain_tl >= 0 else self.LOSS_FONT
            )

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 60

    # -------------------------------------------------------
    # HELPERS
    # -------------------------------------------------------

    def _round(self, value: Decimal) -> Decimal:
        """
        Rounds monetary values to 2 decimal places using ROUND_HALF_UP.

        This ensures regulatory-compliant rounding while keeping
        all arithmetic in Decimal (no float conversion).
        """
        return value.quantize(TWO_PLACES, rounding=ROUND_HALF_UP)

    def _format_money_cell(self, cell) -> None:
        """
        Applies consistent financial formatting:
        - Thousand separator
        - 2 decimal places (visual)
        - Right alignment
        """
        cell.number_format = self.MONEY_FORMAT
        cell.alignment = Alignment(horizontal="right")